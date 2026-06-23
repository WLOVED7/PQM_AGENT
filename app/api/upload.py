"""
Excel 上传接口 - 将 SIP Excel 数据导入 sip_records 表
PDF 上传接口  - 将 SIP PDF 文件保存到 documents/ 目录，以文件号重命名

Excel 列顺序（共12列）:
  1. 项目号  → project
  2. 客户    → customer
  3. 文件号  → document_id
  4. 零件号  → part_num
  5. 零件名称 → part_name
  6. 工序    → process
  7. 模具号  → mold_num
  8. 检验项  → inspection_item
  9. 规范或描述 → specification
  10. 检验方法 → inspection_method
  11. 检查频次 → inspection_frequency
  12. 版本号  → version
"""

import io
import re
from pathlib import Path
from fastapi import APIRouter, File, Form, UploadFile, HTTPException
from openpyxl import load_workbook
from pypdf import PdfReader

from app.db.connection import execute, fetch_one, fetch_all
from app.utils.logger import get_logger

logger = get_logger(__name__)
router = APIRouter(prefix="/upload", tags=["Upload"])

# documents 目录（与 main.py 挂载路径一致）
_DOCUMENTS_DIR = Path(__file__).parent.parent.parent / "documents"

# Excel 列索引（0-based）到字段名的映射
_COL_FIELDS = [
    "project",
    "customer",
    "document_id",
    "part_num",
    "part_name",
    "process",             # 第6列：工序
    "mold_num",
    "inspection_item",
    "specification",
    "inspection_method",
    "inspection_frequency",
    "version",
]

_UPSERT_SQL = """
INSERT INTO sip_records
    (project, customer, document_id, part_num, part_name, process, mold_num,
     inspection_item, specification, inspection_method, inspection_frequency,
     version, chunk_text)
VALUES
    ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13)
ON CONFLICT (document_id, inspection_item) DO UPDATE SET
    project              = EXCLUDED.project,
    customer             = EXCLUDED.customer,
    part_num             = EXCLUDED.part_num,
    part_name            = EXCLUDED.part_name,
    process              = EXCLUDED.process,
    mold_num             = EXCLUDED.mold_num,
    specification        = EXCLUDED.specification,
    inspection_method    = EXCLUDED.inspection_method,
    inspection_frequency = EXCLUDED.inspection_frequency,
    version              = EXCLUDED.version,
    chunk_text           = EXCLUDED.chunk_text,
    updated_at           = CURRENT_TIMESTAMP
RETURNING (xmax = 0) AS inserted
"""


def _cell_str(cell) -> str:
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def _make_chunk_text(row: dict) -> str:
    parts = [
        row.get("inspection_item", ""),
        row.get("specification", ""),
        row.get("inspection_method", ""),
    ]
    return " ".join(p for p in parts if p)


def _extract_document_id(text: str) -> str | None:
    """
    从 PDF 文本中提取文件号，支持多种格式：
    - 策略1: "文件号" / "文件编号" 标签后紧跟的值
    - 策略2: 通用格式 XXXX-SIP/SOP-XXXX-NNN
    - 策略3: 宽松格式 两段以上大写字母+数字用连字符连接
    """
    # 策略1: 文件号/文件编号标签
    m = re.search(r'(?:文件号|文件编号)[：:\s]*([A-Za-z][A-Za-z0-9_-]{3,})', text)
    if m:
        return m.group(1).strip().upper()

    # 策略2: SIP/SOP 通用格式
    m = re.search(r'\b([A-Z]{2,}-(?:SIP|SOP)-[A-Z0-9]+-\d+)\b', text)
    if m:
        return m.group(1)

    # 策略3: 宽松 — 三段或更多，每段2~10位大写字母+数字
    m = re.search(r'\b([A-Z]{2,10}-[A-Z0-9]{2,10}-[A-Z0-9]{2,10}(?:-\d{2,6})?)\b', text)
    if m:
        return m.group(1)

    return None


@router.post("/excel")
async def upload_excel(file: UploadFile = File(...)):
    """
    上传 SIP Excel 文件，将数据批量写入 sip_records 表。
    返回 {inserted, skipped, errors, error_details}
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="只支持 .xlsx / .xls 格式")

    content = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败: {e}")

    inserted = 0
    updated = 0
    skipped = 0
    error_details = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [_cell_str(c) for c in row[:12]]
        if not any(values):
            skipped += 1
            continue

        record = dict(zip(_COL_FIELDS, values))
        missing = [f for f in ("project", "customer", "document_id", "part_num",
                               "part_name", "inspection_item", "version")
                   if not record.get(f)]
        if missing:
            error_details.append(f"第{row_idx}行缺少必填字段: {missing}")
            skipped += 1
            continue

        chunk_text = _make_chunk_text(record)
        params = (
            record["project"],
            record["customer"],
            record["document_id"],
            record["part_num"],
            record["part_name"],
            record["process"] or None,
            record["mold_num"] or None,
            record["inspection_item"],
            record["specification"] or None,
            record["inspection_method"] or None,
            record["inspection_frequency"] or None,
            record["version"],
            chunk_text or None,
        )

        try:
            row_result = await fetch_one(_UPSERT_SQL, params)
            if row_result and row_result["inserted"]:
                inserted += 1
            else:
                updated += 1
        except Exception as e:
            error_details.append(f"第{row_idx}行写入失败: {e}")
            skipped += 1

    logger.info(f"Excel 导入完成: inserted={inserted}, updated={updated}, skipped={skipped}")
    return {
        "success": True,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": len(error_details),
        "error_details": error_details[:20],
    }


@router.post("/pdf")
async def upload_pdf(
    file: UploadFile = File(...),
    document_id: str = Form(None),   # 可选：手动指定文件号
):
    """
    上传 SIP PDF 文件：
    1. 若提供 document_id 参数，直接使用
    2. 否则解析 PDF 文本，自动提取文件号
    3. 以 {文件号}.pdf 保存到 documents/ 目录
    返回 {document_id, filename, overwritten, extracted_by}
    """
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="只支持 .pdf 格式")

    content = await file.read()
    extracted_by = "manual"

    if document_id:
        # 手动指定，做简单格式校验
        document_id = document_id.strip().upper()
        if not re.match(r'^[A-Z][A-Z0-9_-]{2,}$', document_id):
            raise HTTPException(status_code=400, detail=f"文件号格式不合法: {document_id}")
    else:
        # 自动提取
        try:
            reader = PdfReader(io.BytesIO(content))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"PDF 解析失败: {e}")

        document_id = _extract_document_id(text)

        if not document_id:
            # 返回提取到的文本片段，方便前端展示供用户手动填写
            snippet = text[:200].strip().replace("\n", " ") if text.strip() else "（PDF 无可读文本，可能是扫描件）"
            logger.warning(f"PDF 文件号提取失败: {file.filename}, 文本片段: {snippet[:80]}")
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "未能自动识别文件号，请手动输入",
                    "snippet": snippet,
                },
            )
        extracted_by = "auto"

    # 保存文件
    _DOCUMENTS_DIR.mkdir(exist_ok=True)
    dest = _DOCUMENTS_DIR / f"{document_id}.pdf"
    overwritten = dest.exists()
    dest.write_bytes(content)

    logger.info(f"PDF 保存成功: {dest} (overwritten={overwritten}, extracted_by={extracted_by})")
    return {
        "success": True,
        "document_id": document_id,
        "filename": f"{document_id}.pdf",
        "overwritten": overwritten,
        "extracted_by": extracted_by,
    }


@router.get("/customers")
async def get_customers():
    """返回所有客户及其 SIP 统计（供前端首页展示）。"""
    rows = await fetch_all("""
        SELECT
            customer,
            COUNT(DISTINCT document_id)  AS doc_count,
            COUNT(DISTINCT part_name)    AS part_count,
            COUNT(*)                     AS record_count
        FROM sip_records
        GROUP BY customer
        ORDER BY customer
    """)
    return {"customers": rows}


@router.get("/customers/{customer}/parts")
async def get_customer_parts(customer: str):
    """
    返回指定客户下所有零件的 SIP 文件列表。
    每条记录包含 part_name、document_id，并检查 PDF 是否已上传到 documents/ 目录。
    """
    rows = await fetch_all("""
        SELECT
            part_name,
            part_num,
            document_id,
            MAX(version) AS version
        FROM sip_records
        WHERE customer = $1
        GROUP BY part_name, part_num, document_id
        ORDER BY part_name, document_id
    """, (customer,))

    parts = []
    for r in rows:
        doc_id = r["document_id"]
        pdf_path = _DOCUMENTS_DIR / f"{doc_id}.pdf"
        parts.append({
            "part_name":   r["part_name"],
            "part_num":    r["part_num"],
            "document_id": doc_id,
            "version":     r["version"],
            "pdf_exists":  pdf_path.exists(),
            "pdf_url":     f"documents/{doc_id}.pdf" if pdf_path.exists() else None,
        })
    return {"customer": customer, "parts": parts}


@router.get("/prompt-hints")
async def get_prompt_hints():
    """
    从 sip_records 中抽取代表性记录，供前端生成具体的提示词。
    按客户+零件名称+工序+检验项去重，随机取最多 12 条。
    """
    rows = await fetch_all("""
        SELECT DISTINCT ON (customer, part_name, inspection_item)
            customer,
            part_name,
            process,
            inspection_item
        FROM sip_records
        WHERE inspection_item IS NOT NULL
        ORDER BY customer, part_name, inspection_item, random()
        LIMIT 12
    """)
    return {"hints": [dict(r) for r in rows]}
